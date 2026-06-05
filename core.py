"""
job-crawler 공통 로직.

CLI(main.py)와 웹앱(app.py)이 함께 사용하는 핵심 함수 모음.
- 마감일 파싱/필터
- 크롤링 오케스트레이션 (run_crawl)
- 엑셀 생성 (save_excel / df_to_excel_bytes)
"""

import re
from io import BytesIO
from datetime import date, timedelta, datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from crawlers.saramin import SaraminCrawler
from crawlers.wanted import WantedCrawler
from crawlers.jobkorea import JobkoreaCrawler
from crawlers.jumpit import JumpitCrawler
from crawlers.rocketpunch import RocketpunchCrawler

# 사이트 코드 → 크롤러 클래스
CRAWLERS = {
    "wanted": WantedCrawler,
    "jumpit": JumpitCrawler,
    "jobkorea": JobkoreaCrawler,
    "saramin": SaraminCrawler,
    "rocketpunch": RocketpunchCrawler,
}

# 사이트 코드 → 한글 라벨
SITE_LABELS = {
    "wanted": "원티드",
    "jumpit": "점핏",
    "jobkorea": "잡코리아",
    "saramin": "사람인",
    "rocketpunch": "로켓펀치",
}

# 브라우저(Playwright)가 필요한 사이트 — 클라우드 배포 시 불안정할 수 있음
SITE_NEEDS_BROWSER = {"jumpit", "jobkorea", "rocketpunch"}

# 사람인은 API 키가 필요
SITE_NEEDS_KEY = {"saramin"}


# ──────────────────────────────────────────────────────────────
# 마감일 처리
# ──────────────────────────────────────────────────────────────
def parse_deadline(raw: str) -> date | None:
    """마감일 문자열을 date로 변환. 상시/빈값은 None 반환."""
    if not raw or raw.strip() in ("", "상시", "상시채용"):
        return None
    raw = raw.strip()

    m = re.match(r"D-(\d+)", raw)
    if m:
        return date.today() + timedelta(days=int(m.group(1)))

    m = re.match(r"(\d{1,2})/(\d{1,2})", raw)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        year = date.today().year
        try:
            d = date(year, month, day)
            if d < date.today():
                d = date(year + 1, month, day)
            return d
        except ValueError:
            return None

    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", raw)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None

    return None


def is_within_days(raw: str, days: int = 14) -> bool:
    """마감일이 오늘~N일 이내이면 True. 상시/파싱 불가 = 포함."""
    d = parse_deadline(raw)
    if d is None:
        return True
    today = date.today()
    return today <= d <= today + timedelta(days=days)


# ──────────────────────────────────────────────────────────────
# 크롤링 오케스트레이션
# ──────────────────────────────────────────────────────────────
def run_crawl(
    keywords: list[str],
    sites: list[str],
    days: int = 14,
    sort: str = "deadline",
    progress=None,
) -> pd.DataFrame:
    """
    여러 사이트를 순회하며 공고를 수집·정제하여 DataFrame 반환.

    progress: 선택적 콜백. progress(site_code, status, count=None, error=None)
              status ∈ {"start", "done", "error", "skip"}
    """
    all_dfs = []

    for name in CRAWLERS:
        if name not in sites:
            continue

        if progress:
            progress(name, "start")

        try:
            df = CRAWLERS[name]().run(keywords)
            if progress:
                progress(name, "done", count=len(df))
            all_dfs.append(df)
        except Exception as e:  # 한 사이트 실패가 전체를 멈추지 않도록
            if progress:
                progress(name, "error", error=str(e))

    if not all_dfs:
        return pd.DataFrame(
            columns=["title", "company", "location", "experience",
                     "deadline", "url", "source", "collected_at"]
        )

    result = pd.concat(all_dfs, ignore_index=True)
    result = result.drop_duplicates(subset=["url"])

    # 마감 필터
    result = result[result["deadline"].apply(
        lambda d: is_within_days(str(d) if pd.notna(d) else "", days=days)
    )]

    # 정렬
    if sort == "deadline":
        result["_dl"] = result["deadline"].apply(
            lambda d: parse_deadline(str(d) if pd.notna(d) else "") or date(9999, 12, 31)
        )
        result = result.sort_values("_dl").drop(columns=["_dl"]).reset_index(drop=True)
    elif sort == "company":
        result = result.sort_values("company").reset_index(drop=True)
    else:
        result = result.sort_values("source").reset_index(drop=True)

    return result


# ──────────────────────────────────────────────────────────────
# 엑셀 생성
# ──────────────────────────────────────────────────────────────
def _build_workbook(df: pd.DataFrame) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공고 목록"

    headers = ["공고 제목", "회사", "위치", "경력", "마감일", "출처", "수집일"]
    col_widths = [50, 20, 18, 15, 12, 10, 16]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    link_font = Font(color="0563C1", underline="single", size=10)
    normal_font = Font(size=10)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    for i, row in df.reset_index(drop=True).iterrows():
        r = i + 2
        fill = alt_fill if i % 2 == 0 else PatternFill()

        cell = ws.cell(row=r, column=1, value=row["title"])
        if row.get("url"):
            cell.hyperlink = row["url"]
            cell.font = link_font
        else:
            cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = fill
        cell.border = border

        for col, field in enumerate(
            ["company", "location", "experience", "deadline", "source", "collected_at"], 2
        ):
            c = ws.cell(row=r, column=col, value=row.get(field, ""))
            c.font = normal_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = border

        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return wb


def save_excel(df: pd.DataFrame, path: Path) -> None:
    """DataFrame을 서식 적용된 엑셀 파일로 저장."""
    _build_workbook(df).save(path)


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """웹 다운로드용: 엑셀을 메모리(bytes)로 생성."""
    buf = BytesIO()
    _build_workbook(df).save(buf)
    return buf.getvalue()
