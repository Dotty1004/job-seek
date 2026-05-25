import sys
import re
import argparse
from datetime import date, timedelta, datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from crawlers.saramin import SaraminCrawler
from crawlers.wanted import WantedCrawler
from crawlers.jobkorea import JobkoreaCrawler
from crawlers.jumpit import JumpitCrawler
from crawlers.rocketpunch import RocketpunchCrawler

CRAWLERS = {
    "saramin": SaraminCrawler,
    "wanted": WantedCrawler,
    "jobkorea": JobkoreaCrawler,
    "jumpit": JumpitCrawler,
    "rocketpunch": RocketpunchCrawler,
}


def parse_deadline(raw: str) -> date | None:
    """마감일 문자열을 date로 변환. 상시/빈값은 None 반환."""
    if not raw or raw.strip() in ("", "상시", "상시채용"):
        return None
    raw = raw.strip()

    m = re.match(r"D-(\d+)", raw)
    if m:
        return date.today() + timedelta(days=int(m.group(1)))

    m = re.match(r"(\d{1,2})/(\d{1,2})", raw)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        year = date.today().year
        try:
            d = date(year, month, day)
            if d < date.today():
                d = date(year + 1, month, day)
            return d
        except ValueError:
            return None

    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", raw)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None

    return None


def is_within_days(raw: str, days: int = 14) -> bool:
    """마감일이 오늘~N일 이내이면 True. 상시/파싱 불가 = 포함."""
    d = parse_deadline(raw)
    if d is None:
        return True
    today = date.today()
    return today <= d <= today + timedelta(days=days)


def save_excel(df: pd.DataFrame, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공고 목록"

    headers = ["공고 제목", "회사", "위치", "경력", "마감일", "출처", "수집일"]
    col_widths = [50, 20, 18, 15, 12, 10, 16]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    link_font = Font(color="0563C1", underline="single", size=10)
    normal_font = Font(size=10)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    for i, row in df.iterrows():
        r = i + 2
        fill = alt_fill if i % 2 == 0 else PatternFill()

        cell = ws.cell(row=r, column=1, value=row["title"])
        if row.get("url"):
            cell.hyperlink = row["url"]
            cell.font = link_font
        else:
            cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = fill
        cell.border = border

        for col, field in enumerate(["company", "location", "experience", "deadline", "source", "collected_at"], 2):
            c = ws.cell(row=r, column=col, value=row.get(field, ""))
            c.font = normal_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = border

        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(path)


def parse_args():
    parser = argparse.ArgumentParser(description="job-crawler: 채용공고 수집기")
    parser.add_argument(
        "--keywords", type=str, default=None,
        help="콤마 구분 키워드 (예: AX,LLM). 미지정 시 config.KEYWORDS 사용",
    )
    parser.add_argument(
        "--days", type=int, default=14,
        help="마감일 필터: n일 이내 공고만 수집 (기본값: 14, 365=전체)",
    )
    parser.add_argument(
        "--sites", type=str, default=None,
        help="실행할 사이트 콤마 구분 (예: wanted,jumpit). 미지정 시 config.SITES 사용",
    )
    parser.add_argument(
        "--sort", type=str, default="deadline",
        choices=["deadline", "company", "source"],
        help="정렬 기준: deadline(마감일 빠른 순), company(회사명), source(사이트별)",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    keywords = [k.strip() for k in args.keywords.split(",")] if args.keywords else config.KEYWORDS
    active_sites = (
        {s.strip() for s in args.sites.split(",")} if args.sites
        else {name for name, on in config.SITES.items() if on}
    )

    print(f"키워드: {keywords}")
    print(f"사이트: {sorted(active_sites)}")
    print(f"마감 필터: {'전체 (마감 지난 공고만 제외)' if args.days >= 365 else f'{args.days}일 이내'}")
    print(f"정렬: {args.sort}")

    all_dfs = []

    for name, cls in CRAWLERS.items():
        if name not in active_sites:
            print(f"[{name}] 건너뜀")
            continue

        print(f"\n{'='*40}")
        print(f"[{name}] 크롤링 시작 - 키워드 {len(keywords)}개")
        crawler = cls()
        df = crawler.run(keywords)
        print(f"[{name}] 수집 완료: {len(df)}건")
        all_dfs.append(df)

    if not all_dfs:
        print("\n수집된 데이터가 없습니다.")
        sys.exit(0)

    result = pd.concat(all_dfs, ignore_index=True)
    result = result.drop_duplicates(subset=["url"])

    before = len(result)
    result = result[result["deadline"].apply(
        lambda d: is_within_days(str(d) if pd.notna(d) else "", days=args.days)
    )]

    if args.days >= 365:
        print(f"\n기간 제한 없음 (마감 지난 공고만 제외): {before - len(result)}건")
    else:
        print(f"\n{args.days}일 초과 마감 제외: {before - len(result)}건")

    # 정렬
    if args.sort == "deadline":
        result["_dl"] = result["deadline"].apply(
            lambda d: parse_deadline(str(d) if pd.notna(d) else "") or date(9999, 12, 31)
        )
        result = result.sort_values("_dl").drop(columns=["_dl"]).reset_index(drop=True)
    elif args.sort == "company":
        result = result.sort_values("company").reset_index(drop=True)
    else:
        result = result.sort_values("source").reset_index(drop=True)

    out_dir = Path(__file__).parent / "output"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"jobs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    save_excel(result, out_path)

    print(f"\n저장 완료: {out_path}")
    print(f"총 {len(result)}건 (사이트별: {result.groupby('source').size().to_dict()})")


if __name__ == "__main__":
    main()
